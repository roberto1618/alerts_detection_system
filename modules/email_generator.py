"""
This module uses the final alerts table summary to generate an email with the alerts information.
"""

from datetime import datetime
from google.cloud import bigquery
from config.alerts_table_styles.extra_styles import ExtraStylesAlerts
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email.utils import formataddr
import logging
from premailer import transform
import pandas as pd
from email import encoders
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.drawing.image import Image
import os

class EmailGenerator():
    """
    This class creates the email corpus, sets the smtp service with email and sends the email with the alerts.
    """

    def __init__(self, email_from, emails_to, password, smtp_server, port):
        """
        Init parameters

        :param email_from: The email sender.
        :type email_from: string
        :param emails_to: The recipients of the email.
        :type emails_to: list
        :param password: The password to login in the sender account.
        :type password: string
        :param smtp_server: The smtp server
        :type smtp_server: string
        :param port: The port to establish the connection
        :type port: integer
        """

        self.email_from = email_from
        self.emails_to = emails_to
        self.password = password
        self.port = port
        self.smtp_server = smtp_server

    def create_corpus(self, only_alerts_table, date, alerts_table_title, future_table = None):
        """
        Function to create the corpus of the email. If any alert is detected, the corpus is created with a message and
        a table with the metrics that have alerts.

        :param alerts_table: The table that has the summary of every analyzed metric and if they have an alert or not.
        :type alerts_table: Pandas DataFrame
        :param future_table: Table with future predictions
        :type future_table: Pandas DataFrame
        :param alerts_table_title: The title of the alerts table in the email.
        :type alerts_table_title: string
        :param date: The date when the alerts System is evaluating the alert.
        :type date: string
        """

        # The message of the email has different parts
        msg = MIMEMultipart()
        msg['Subject'] = 'Alertas'
        msg['From'] = formataddr(('Alertas EAM', self.email_from))
        msg['To'] = self.emails_to

        # Depending on an alert is detected or not, a different message is sent.
        # If the sum of the column 'Alert' of the general alert table info is greater than 0, at least one alert is detected
        alert_exist = only_alerts_table.shape[0] > 0
        if alert_exist:
            # Create the text and alerts table and convert them to html
            text = 'Estimated customer, <br><br> We have detected some alerts in the following \
                       metrics on <b>' + date.strftime('%Y-%m-%d') + '</b>. \
                       Please revise them to understand the issue or fix the problem.'

            text = f'<p style="font-size: 16px;">{text}</p>'

            # Add css styles
            with open('./config/alerts_table_styles/style.css', 'r') as f:
                styles = f.read()

            # Add the logo to the alerts table. You must save the logo in config/alerts_table_styles as a png image
            try:
                with open('./config/alerts_table_styles/logo.png', 'rb') as f:
                    img_data = f.read()
                    logo = MIMEImage(img_data)
                    logo.add_header('Content-ID', '<logo>')
                    logo.add_header("Content-Disposition", "inline; filename=logo.png")
                    msg.attach(logo)
                
                logo_html_part = '<img src="cid:logo" alt="Logotipo" class="logo" />'
            except FileNotFoundError:
                logo_html_part = ''
            
            # Convert the table to html
            # The extra styles will be applied whenever new methods in ExtraStylesAlerts are created
            extra_styles_inst = ExtraStylesAlerts()
            table_html_fin = ''
            extra_styles_methods = [name_method for name_method in dir(extra_styles_inst) if callable(getattr(extra_styles_inst, name_method)) and not '__' in name_method]
            if len(extra_styles_methods) > 1:
                extra_styles_methods.remove('do_nothing')
            for name_method in extra_styles_methods:
                method = getattr(extra_styles_inst, name_method)
                table_html = method(only_alerts_table)
                table_html_fin = table_html_fin + table_html + '<br>'

            # Set the title of the alerts table
            title_text = alerts_table_title

            # Create the header with the logo and title
            header_html = f"""
            <div class="table-container">
            <div class="header">
                {logo_html_part}
                <h1 class="title">{title_text}</h1>
            </div>
            """

            # Join header and table
            table_html = header_html + table_html_fin + "</div>"

            # Apply the css styles in line
            table_html_with_styles = f"<style>{styles}</style>{table_html}"

            # Convert the CSS styles into line styles
            table_html = transform(table_html_with_styles)

            # Add the alerts details below the table
            if only_alerts_table['Details'].isin(['Decreasing tendency.*', 'Increasing tendency.*']).any():
                details_tendency = '*It is possible that your yesterday\'s data should be higher or lower given the historical behavior of the metric. Please check if your data is changing the tendency.'
            else:
                details_tendency = ''

            details_tendency = f'<p style="font-size: 13px;">{details_tendency}</p>'

            if only_alerts_table['Details'].isin(['Missing data.**']).any():
                details_missing = '**Your yesterday\'s data is missed. Please check it there is a problem in the BQ extraction.'
            else:
                details_missing = ''
            
            details_missing = f'<p style="font-size: 13px;">{details_missing}</p>'

            if only_alerts_table['Details'].isin(['Constraint violated.***']).any():
                details_constraint = '***The constraint in the Metric column is violated. Please check your metrics.'
            else:
                details_constraint = ''

            details_constraint = f'<p style="font-size: 13px;">{details_constraint}</p>'

            # Join all elements to create the corpus
            corpus = MIMEText(text + table_html + details_tendency + details_missing + details_constraint, 'html')

            msg.attach(corpus)

        else:
            # If there is no alerts, send this message
            text = 'Estimated customer, you do not have any alerts on <b>' + date.strftime('%Y-%m-%d') + '<b>.'
            text = f'<p style="font-size: 16px;">{text}</p>'
            text_part = MIMEText(text, 'html')

            msg.attach(text_part)

        # ## Attach predictions if they exist
        # if future_table is not None:
        #     # Write the excel as a workbook
        #     excel_archive = pd.ExcelWriter('future_table.xlsx', engine = 'openpyxl')
        #     workbook = excel_archive.book
        #     worksheet = workbook.create_sheet('Future Table')
        #     # Edit the font of the columns and other style things
        #     font = Font(bold = True, color = 'FFFFFF')
        #     fill = PatternFill(start_color = 'eb5959', end_color = 'eb5959', fill_type = 'solid')
        #     alignment = Alignment(horizontal = 'center')
        #     # Add every row of the future table to the excel archive
        #     for row in dataframe_to_rows(future_table, index = False, header = True):
        #         worksheet.append(row)

        #     # Apply style to the cells. The header will have a different style
        #     last_col = None
        #     for col in worksheet.iter_cols():
        #         column = col[0].column_letter
        #         max_length = 0
        #         for cell in col:
        #             if isinstance(cell.value, str):
        #                 try:
        #                     if len(str(cell.value)) > max_length:
        #                         max_length = len(cell.value)
        #                 except:
        #                     pass

        #                 cell.font = font
        #                 cell.fill = fill
        #                 cell.alignment = alignment
                    
        #             last_col = cell.column_letter
        #         # Adjust the width of the columns
        #         adjusted_width = (max_length + 2) * 1.2
        #         worksheet.column_dimensions[column].width = adjusted_width
            
        #     # The first column with the dates will have a different colour
        #     for row in worksheet.iter_rows(min_row = 2):
        #         for cell in row:
        #             if cell.column == 1 and cell.row != 1:
        #                 cell.font = Font(bold = False, color = 'FFFFFF')
        #                 cell.fill = PatternFill(start_color = 'eb5959', end_color = 'eb5959', fill_type = 'solid')

        #     # Attach the logo to the end of the table
        #     try:
        #         logo = Image('./config/alerts_table_styles/logo.png')

        #         # Set the logo width and height
        #         logo.width = 18.5
        #         logo.height = 18.5

        #         # Get the last column in order to set the logo in that place
        #         if last_col:
        #             next_col_index = column_index_from_string(last_col) + 1
        #             next_col = get_column_letter(next_col_index)

        #         if next_col:
        #             last_row = worksheet.max_row
        #             next_cell = worksheet[next_col + '1']
        #             worksheet.add_image(logo, next_cell.coordinate)

        #     except FileNotFoundError:
        #         logo = ''

        #     excel_archive.save()

        #     # Read the excel and attach to the email with a title
        #     with open('future_table.xlsx', 'rb') as f:
        #         file_attach = MIMEBase('application', 'octet-stream')
        #         file_attach.set_payload(f.read())
        #         encoders.encode_base64(file_attach)
        #         file_attach.add_header('Content-Disposition', 'attachment', filename = 'Predictions from ' + date.strftime('%Y-%m-%d') + '.xlsx')
        #         msg.attach(file_attach)
            
        #     # Remove the excel from the directory of the project
        #     os.remove('future_table.xlsx')

        return msg
    
    def set_email_service(self):
        """
        Function to set the connection with email service.
        """

        # Create the connection with ssl and smtp protocol
        # The credentials and email info from the config are used
        context = ssl.create_default_context()

        server = smtplib.SMTP(self.smtp_server, self.port)
        server.starttls(context = context)
        server.login(self.email_from, self.password)

        return server
    
    def send_email(self, msg):
        """
        Function to send the email to the recipients.

        :param msg: The message to send.
        :type msg: string
        """

        server = self.set_email_service()
        server.send_message(msg)

